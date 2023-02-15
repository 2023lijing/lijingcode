#  单元格写入样式

from openpyxl.styles import PatternFill, Font


# pass样式
def pass_(cell, row, column):
    # 写入内容
    cell(row=row, column=column).value = 'Pass'
    # 单元格样式定义：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
    cell(row=row, column=column).font = Font(bold=True)


# Failed样式
def failed(cell, row, column):
    # 写入内容
    cell(row=row, column=column).value = 'Failed'
    # 单元格样式定义：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
    cell(row=row, column=column).font = Font(bold=True)
