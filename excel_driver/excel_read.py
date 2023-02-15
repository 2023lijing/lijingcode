# excel 文件读取类，用于实现测试用例文件的读取与执行

import openpyxl as openpyxl

# 解析测试用例中测试参数单元格的内容，并转换为字典的形态返回；
from openpyxl import load_workbook
from selenium.webdriver import Keys
from conf import log_conf
from excel_driver import excel_conf
from web_keys.keys import Keyword


# def arguments(value):
#     data = dict()
#     if value:
#         test_temp = value.split(';')
#         for temp in test_temp:
#             t = temp.split('=', 1)
#             data[t[0]] = t[1]
#     else:
#         pass
#     return data


def run(file_name, log):
    # log = log_conf.get_log('../conf/log.ini')
    # 指定excel路径
    excel = openpyxl.load_workbook(file_name)
    # 获取所有的sheet页，来执行里边的测试内容
    # sheet = excel['Sheet1']
    try:
        # 循环每个sheet页，进行读取
        for name in excel.sheetnames:
            sheet = excel[name]
            log.info('***正在执行{}Sheet页***'.format(name))
            # 获取测试用例正文的内容
            # 读取到所有的表单数据
            for values in sheet.values:
                if type(values[0]) is int:
                    # 接收每一行操作行为对应的参数内容
                    data = {}
                    data['name'] = values[2]
                    data['value'] = values[3]
                    data['txt'] = values[4]
                    data['expected'] = values[6]
                    # 清除参数字典中为None的参数键值对
                    for key in list(data.keys()):
                        if data[key] is None:
                            del data[key]
                    log.info('当前正在执行:{}'.format(values[5]))
                    # print('****正在执行：{}****'.format(values[3]))
                    # data = arguments(values[2])
                    # print(data)
                    # 操作行为的调用分为以下几种类型：
                    # 1、实例化
                    # 2、基于实例化对象进行的操作行为
                    # 3、断言机制，有预期与实际的对比，以及有单元格测试结果的写入；
                    # 实例化操作
                    if values[1] == 'open_browser':
                        keys = Keyword(data['txt'], log)
                    # 常规操作行为
                    # 断言行为：基于断言的返回结果，判断测试的成功与失败；并进行写入操作
                    elif 'assert' in values[1]:
                        status = getattr(keys, values[1])(**data)
                        if status:
                            excel_conf.pass_(sheet.cell, row=values[0] + 2, column=8)
                        else:
                            excel_conf.failed(sheet.cell, row=values[0] + 2, column=8)
                        # 保存excel:确保每次写入都可以被保存
                        excel.save(file_name)
                    else:
                        getattr(keys, values[1])(**data)
    except Exception as e:
        log.exception('运行异常:{}'.format(e))
    finally:
        excel.close()



